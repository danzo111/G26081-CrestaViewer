"""Export the current network.json to a review workbook: Manholes + Pipes."""
import json, math
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NET = r"C:\Users\User\Documents\GitHub\G26081 CrestaViwer\data\network.json"
OUT = r"C:\Users\User\Downloads\G26081_Current_Network_Audit.xlsx"

net = json.load(open(NET))
M = {m["id"]: m for m in net["manholes"]}
pipes = net["pipes"]

def isd(mid):
    m = M.get(mid)
    return mid.startswith("DUMMY") or (m and m.get("type") == "Dummy")

deg = defaultdict(int)
for p in pipes:
    deg[p["from_mh"]] += 1; deg[p["to_mh"]] += 1

def inv(mid, d):
    m = M.get(mid)
    if not m or m.get("cover_elev") is None: return None
    return round(m["cover_elev"] - (d or 0), 3)

wb = openpyxl.Workbook()
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(color="FFFFFF", bold=True)

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"

# ---- Manholes ----
ws = wb.active; ws.title = "Manholes"
mh_cols = ["ID","Type","Is Dummy","X (net)","Y (net)","Cover Level","Depth",
           "Invert Level","Connections","Parent MH"]
ws.append(mh_cols)
for m in sorted(net["manholes"], key=lambda m: m["id"]):
    cov = m.get("cover_elev"); dep = m.get("depth")
    il = None if cov is None else round(cov - (dep or 0), 3)
    ws.append([m["id"], m.get("type",""), "Y" if isd(m["id"]) else "",
               m.get("x"), m.get("y"), cov, dep, il, deg[m["id"]], m.get("parent_mh","")])
style_header(ws, len(mh_cols))

# ---- Pipes ----
ws2 = wb.create_sheet("Pipes")
p_cols = ["ID","From MH","To MH","Type","Diameter (mm)","Dia Source",
          "From Depth","To Depth","From Invert","To Invert","Fall (m)",
          "Length (m)","Flow Override","Touches Dummy"]
ws2.append(p_cols)
for p in sorted(pipes, key=lambda p: p["id"]):
    f, t = p["from_mh"], p["to_mh"]
    fi, ti = inv(f, p.get("from_depth")), inv(t, p.get("to_depth"))
    fall = None if (fi is None or ti is None) else round(fi - ti, 3)
    fm, tm = M.get(f), M.get(t)
    length = None
    if fm and tm and fm.get("x") is not None and tm.get("x") is not None:
        length = round(math.hypot(fm["x"]-tm["x"], fm["y"]-tm["y"]), 2)
    ptype = p.get("type","") or ("Stormwater" if (fm and fm.get("type")=="Stormwater") or (tm and tm.get("type")=="Stormwater") else "Sewer")
    ws2.append([p["id"], f, t, ptype, p.get("diameter_mm"), p.get("diameter_source",""),
                p.get("from_depth"), p.get("to_depth"), fi, ti, fall, length,
                "Y" if p.get("flow_override") else "", "Y" if (isd(f) or isd(t)) else ""])
style_header(ws2, len(p_cols))

# column widths
for ws_ in (ws, ws2):
    for c in range(1, ws_.max_column+1):
        ws_.column_dimensions[get_column_letter(c)].width = 14

wb.save(OUT)

# ---- console summary ----
reals = [m for m in net["manholes"] if not isd(m["id"])]
dums = [m for m in net["manholes"] if isd(m["id"])]
print(f"WROTE {OUT}")
print(f"\nMANHOLES: {len(net['manholes'])} ({len(reals)} real, {len(dums)} dummy)")
from collections import Counter
print("  real by type:", dict(Counter(m.get('type') for m in reals)))
print(f"  with cover level: {sum(1 for m in net['manholes'] if m.get('cover_elev') is not None)}")
print(f"  with depth>0:     {sum(1 for m in net['manholes'] if (m.get('depth') or 0)>0)}")
print(f"  isolated (no pipes): real {sum(1 for m in reals if deg[m['id']]==0)}, dummy {sum(1 for m in dums if deg[m['id']]==0)}")
print(f"\nPIPES: {len(pipes)}")
print("  by type:", dict(Counter((p.get('type') or 'inferred') for p in pipes)))
print(f"  with diameter: {sum(1 for p in pipes if p.get('diameter_mm'))}/{len(pipes)}")
print(f"  flow_override: {sum(1 for p in pipes if p.get('flow_override'))}")
print(f"  touching a dummy: {sum(1 for p in pipes if isd(p['from_mh']) or isd(p['to_mh']))}")
# coordinate extent
xs=[m['x'] for m in reals]; ys=[m['y'] for m in reals]
print(f"\nCoordinate extent (net frame): X[{min(xs):.1f}, {max(xs):.1f}]  Y[{min(ys):.1f}, {max(ys):.1f}]")
